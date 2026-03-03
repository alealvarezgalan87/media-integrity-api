"""Generate an Excel evaluation report from audit data."""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Styles ──────────────────────────────────────────────────────────────────

DARK_BG = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
HEADER_BG = PatternFill(start_color="2C3E6B", end_color="2C3E6B", fill_type="solid")
DOMAIN_BG = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
GREEN_BG = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
LIGHT_GREEN_BG = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
YELLOW_BG = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
ORANGE_BG = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
RED_BG = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
PASS_BG = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
FAIL_BG = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
LIGHT_GRAY_BG = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")
WHITE_BG = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
WHITE_FONT_SM = Font(color="FFFFFF", size=10)
DARK_FONT = Font(color="1B2A4A", size=10)
DARK_FONT_BOLD = Font(color="1B2A4A", bold=True, size=10)
TITLE_FONT = Font(color="FFFFFF", bold=True, size=16)
SUBTITLE_FONT = Font(color="FFFFFF", bold=True, size=12)

SEVERITY_FONTS = {
    "critical": Font(color="FFFFFF", bold=True, size=10),
    "high": Font(color="FFFFFF", bold=True, size=10),
    "medium": Font(color="1B2A4A", bold=True, size=10),
    "low": Font(color="1B2A4A", size=10),
}
SEVERITY_FILLS = {
    "critical": RED_BG,
    "high": ORANGE_BG,
    "medium": YELLOW_BG,
    "low": LIGHT_GREEN_BG,
}

THIN_BORDER = Border(
    left=Side(style="thin", color="BDC3C7"),
    right=Side(style="thin", color="BDC3C7"),
    top=Side(style="thin", color="BDC3C7"),
    bottom=Side(style="thin", color="BDC3C7"),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

DOMAIN_LABELS = {
    "demand_capture_integrity": "Demand Capture Integrity",
    "automation_exposure": "Automation Exposure",
    "measurement_integrity": "Measurement Integrity",
    "capital_allocation_discipline": "Capital Allocation Discipline",
    "creative_velocity": "Creative Velocity",
}

DOMAIN_WEIGHTS = {
    "demand_capture_integrity": 25,
    "automation_exposure": 20,
    "measurement_integrity": 25,
    "capital_allocation_discipline": 20,
    "creative_velocity": 10,
}

IMPL_DESCRIPTIONS = {
    "REDUCE": "Decrease total media spend until structural issues are resolved.",
    "REWEIGHT": "Total spend level may be appropriate but allocation is wrong. Shift budget.",
    "TEST": "Structure is mostly sound but specific hypotheses need validation before scaling.",
    "HOLD": "Maintain current trajectory. Structure is healthy. Focus on incremental optimization.",
}

DOMAIN_ORDER = (
    "demand_capture_integrity",
    "automation_exposure",
    "measurement_integrity",
    "capital_allocation_discipline",
    "creative_velocity",
)


def _risk_band(score: int) -> tuple[str, PatternFill]:
    if score >= 91:
        return "Excellent", GREEN_BG
    elif score >= 76:
        return "Sound", LIGHT_GREEN_BG
    elif score >= 61:
        return "Moderate", YELLOW_BG
    elif score >= 41:
        return "High Exposure", ORANGE_BG
    else:
        return "Critical Failure", RED_BG


def _apply_row_fill(ws, row: int, col_start: int, col_end: int, fill, exclude=None):
    exclude = exclude or set()
    for col in range(col_start, col_end + 1):
        if col not in exclude:
            ws.cell(row=row, column=col).fill = fill


def _apply_row_border(ws, row: int, col_start: int, col_end: int):
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).border = THIN_BORDER


def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _title_row(ws, row: int, text: str, col_end: int, font=TITLE_FONT, fill=DARK_BG):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=1, value=text)
    c.font = font
    c.fill = fill
    c.alignment = CENTER
    for col in range(1, col_end + 1):
        ws.cell(row=row, column=col).fill = fill


def _header_row(ws, row: int, headers: list[str], fill=HEADER_BG):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = WHITE_FONT
        c.fill = fill
        c.alignment = CENTER
        c.border = THIN_BORDER


def _build_advanced_health_sheet(wb, raw_data: dict, account_name: str, date_str: str):
    """Build the Advanced Health sheet with Phase 3 metrics.

    Sections: Quality Score, Negative Keywords, Shopping Structure,
    PMax Audiences, Customer Lists, NCA Settings, GA4 Events.
    """
    has_any = any(raw_data.get(k) for k in (
        "keyword_quality_score", "negative_keywords", "shopping_structure",
        "pmax_audience_signals", "customer_lists", "nca_settings",
    ))
    if not has_any:
        return

    ws = wb.create_sheet("Advanced Health")
    ws.sheet_properties.tabColor = "8E44AD"
    _set_col_widths(ws, [35, 20, 20, 20, 25])

    _title_row(ws, 1, f"Advanced Health Analysis — {account_name}", 5)
    _title_row(ws, 2, f"Phase 3 diagnostic metrics  |  Period: {date_str}",
               5, font=WHITE_FONT_SM, fill=HEADER_BG)

    row = 4

    # ── Quality Score Breakdown ──────────────────────────────────
    qs_data = raw_data.get("keyword_quality_score", [])
    if qs_data:
        _title_row(ws, row, "Quality Score Analysis", 5, font=SUBTITLE_FONT, fill=DOMAIN_BG)
        row += 1
        _header_row(ws, row, ["Keyword", "Campaign", "Quality Score", "Impressions", "Classification"])
        row += 1

        from engine.normalization.brand_classifier import _classify_campaign
        brand_name = raw_data.get("_brand_name", "")

        for r in qs_data[:100]:
            criterion = r.get("adGroupCriterion", r.get("ad_group_criterion", {}))
            if not isinstance(criterion, dict):
                continue
            qi = criterion.get("qualityInfo", criterion.get("quality_info", {}))
            kw_info = criterion.get("keyword", {})
            campaign = r.get("campaign", {})
            metrics = r.get("metrics", {})

            qs_val = qi.get("qualityScore", qi.get("quality_score"))
            if qs_val is None:
                continue

            kw_text = kw_info.get("text", "")
            camp_name = campaign.get("name", "")
            impressions = int(metrics.get("impressions", 0) or 0)

            try:
                classification = _classify_campaign(camp_name, brand_name)
            except Exception:
                classification = "unknown"

            ws.cell(row=row, column=1, value=kw_text).font = DARK_FONT
            ws.cell(row=row, column=1).alignment = LEFT_WRAP
            ws.cell(row=row, column=2, value=camp_name).font = DARK_FONT
            ws.cell(row=row, column=2).alignment = LEFT_WRAP
            c = ws.cell(row=row, column=3, value=int(qs_val))
            c.font = DARK_FONT_BOLD
            c.alignment = CENTER
            if int(qs_val) >= 7:
                c.fill = PASS_BG
            elif int(qs_val) <= 4:
                c.fill = FAIL_BG
            ws.cell(row=row, column=4, value=impressions).font = DARK_FONT
            ws.cell(row=row, column=4).alignment = CENTER
            ws.cell(row=row, column=4).number_format = '#,##0'
            ws.cell(row=row, column=5, value=classification.title()).font = DARK_FONT
            ws.cell(row=row, column=5).alignment = CENTER
            _apply_row_border(ws, row, 1, 5)
            row += 1

        row += 1

    # ── Negative Keywords Summary ────────────────────────────────
    nk_data = raw_data.get("negative_keywords", [])
    if nk_data:
        inner = nk_data[0] if nk_data and isinstance(nk_data[0], dict) else {}
        camp_negatives = inner.get("campaign_negatives", [])

        _title_row(ws, row, "Negative Keywords Summary", 5, font=SUBTITLE_FONT, fill=DOMAIN_BG)
        row += 1
        _header_row(ws, row, ["Campaign", "Channel", "Negative Keyword", "Match Type", ""])
        row += 1

        for r in camp_negatives[:100]:
            campaign = r.get("campaign", {}) if isinstance(r.get("campaign"), dict) else {}
            criterion = r.get("campaign_criterion", r.get("campaignCriterion", {}))
            if isinstance(criterion, dict):
                kw_info = criterion.get("keyword", {})
            else:
                kw_info = {}

            ws.cell(row=row, column=1, value=campaign.get("name", "")).font = DARK_FONT
            ws.cell(row=row, column=1).alignment = LEFT_WRAP
            ws.cell(row=row, column=2, value=campaign.get("advertising_channel_type", "")).font = DARK_FONT
            ws.cell(row=row, column=2).alignment = CENTER
            ws.cell(row=row, column=3, value=kw_info.get("text", "")).font = DARK_FONT
            ws.cell(row=row, column=3).alignment = LEFT_WRAP
            ws.cell(row=row, column=4, value=kw_info.get("matchType", kw_info.get("match_type", ""))).font = DARK_FONT
            ws.cell(row=row, column=4).alignment = CENTER
            _apply_row_border(ws, row, 1, 5)
            row += 1

        row += 1

    # ── Shopping Structure ───────────────────────────────────────
    ss_data = raw_data.get("shopping_structure", [])
    if ss_data:
        inner = ss_data[0] if ss_data and isinstance(ss_data[0], dict) else {}
        product_groups = inner.get("product_groups", [])
        campaign_audiences = inner.get("campaign_audiences", [])

        if product_groups or campaign_audiences:
            _title_row(ws, row, "Shopping Structure", 5, font=SUBTITLE_FONT, fill=DOMAIN_BG)
            row += 1

            # Summary metrics
            from engine.normalization.shopping_structure import compute_shopping_structure_metrics
            ss_metrics = compute_shopping_structure_metrics(ss_data)

            summary_items = [
                ("Product Overlap %", f"{ss_metrics['shopping_campaign_product_overlap_pct'] * 100:.1f}%"),
                ("Shopping RLSA Campaigns", str(ss_metrics['shopping_rlsa_campaign_count'])),
                ("Product Groups Analyzed", str(len(product_groups))),
                ("Campaign Audiences", str(len(campaign_audiences))),
            ]
            for label, val in summary_items:
                ws.cell(row=row, column=1, value=label).font = DARK_FONT_BOLD
                ws.cell(row=row, column=1).alignment = LEFT_WRAP
                ws.cell(row=row, column=2, value=val).font = DARK_FONT_BOLD
                ws.cell(row=row, column=2).alignment = CENTER
                _apply_row_border(ws, row, 1, 2)
                row += 1

            row += 1

    # ── PMax Audience Signals ────────────────────────────────────
    pmax_data = raw_data.get("pmax_audience_signals", [])
    if pmax_data:
        from engine.normalization.pmax_audiences import compute_pmax_audience_metrics
        pmax_metrics = compute_pmax_audience_metrics(pmax_data)

        _title_row(ws, row, "PMax Audience Signals", 5, font=SUBTITLE_FONT, fill=DOMAIN_BG)
        row += 1
        summary_items = [
            ("Prospecting Campaigns", str(pmax_metrics['pmax_prospecting_campaign_count'])),
            ("Retargeting Campaigns", str(pmax_metrics['pmax_retargeting_campaign_count'])),
        ]
        split_ok = pmax_metrics['pmax_prospecting_campaign_count'] > 0 and pmax_metrics['pmax_retargeting_campaign_count'] > 0
        summary_items.append(("Pro/Ret Split", "Yes" if split_ok else "No — Missing split"))

        for label, val in summary_items:
            ws.cell(row=row, column=1, value=label).font = DARK_FONT_BOLD
            ws.cell(row=row, column=1).alignment = LEFT_WRAP
            c = ws.cell(row=row, column=2, value=val)
            c.font = DARK_FONT_BOLD
            c.alignment = CENTER
            if "Missing" in val:
                c.fill = FAIL_BG
            elif val in ("Yes",):
                c.fill = PASS_BG
            _apply_row_border(ws, row, 1, 2)
            row += 1

        row += 1

    # ── Customer Lists ───────────────────────────────────────────
    cl_data = raw_data.get("customer_lists", [])
    if cl_data:
        from engine.normalization.customer_lists import compute_customer_list_metrics
        cl_metrics = compute_customer_list_metrics(cl_data, raw_data.get("change_history", []))

        _title_row(ws, row, "Customer Lists Health", 5, font=SUBTITLE_FONT, fill=DOMAIN_BG)
        row += 1

        days = cl_metrics['days_since_customer_list_refresh']
        match_rate = cl_metrics['customer_list_match_rate']

        summary_items = [
            ("Days Since Last Refresh", str(days)),
            ("Average Match Rate", f"{match_rate * 100:.1f}%"),
            ("CRM-Based Lists Found", str(len([r for r in cl_data if isinstance(r.get("userList"), dict) and "CRM" in str(r.get("userList", {}).get("type", ""))]))),
        ]
        for label, val in summary_items:
            ws.cell(row=row, column=1, value=label).font = DARK_FONT_BOLD
            ws.cell(row=row, column=1).alignment = LEFT_WRAP
            c = ws.cell(row=row, column=2, value=val)
            c.font = DARK_FONT_BOLD
            c.alignment = CENTER
            if "days" in label.lower() and days > 90:
                c.fill = FAIL_BG
            elif "rate" in label.lower() and match_rate < 0.30:
                c.fill = FAIL_BG
            _apply_row_border(ws, row, 1, 2)
            row += 1

        row += 1

    # ── NCA Settings ─────────────────────────────────────────────
    nca_data = raw_data.get("nca_settings", [])
    if nca_data:
        from engine.normalization.nca_settings import compute_nca_metrics
        nca_metrics = compute_nca_metrics(nca_data)

        _title_row(ws, row, "New Customer Acquisition Settings", 5, font=SUBTITLE_FONT, fill=DOMAIN_BG)
        row += 1

        bid_adj = nca_metrics['nca_bid_adjustment']
        validated = nca_metrics['nca_bid_validation']

        summary_items = [
            ("NCA Bid Adjustment", f"${bid_adj:.0f}" if bid_adj > 0 else "Not Set"),
            ("Bid Validated", "Yes" if validated else "No — Unverified"),
        ]
        for label, val in summary_items:
            ws.cell(row=row, column=1, value=label).font = DARK_FONT_BOLD
            ws.cell(row=row, column=1).alignment = LEFT_WRAP
            c = ws.cell(row=row, column=2, value=val)
            c.font = DARK_FONT_BOLD
            c.alignment = CENTER
            if "Unverified" in val:
                c.fill = FAIL_BG
            _apply_row_border(ws, row, 1, 2)
            row += 1


def generate_audit_excel(audit_data: dict) -> bytes:
    """Generate an Excel workbook from audit data and return as bytes.

    Args:
        audit_data: Dict with keys: account_name, account_id, date_range,
                    scoring, domain_scores, red_flags, execution, run_id,
                    raw_data (optional — granular extractor output for detailed sheets).

    Returns:
        Excel file content as bytes.
    """
    wb = Workbook()

    account_name = audit_data.get("account_name", "Unknown")
    account_id = audit_data.get("account_id", "")
    date_range = audit_data.get("date_range", {})
    scoring = audit_data.get("scoring", {})
    domain_scores = audit_data.get("domain_scores", {})
    red_flags = audit_data.get("red_flags", [])
    run_id = audit_data.get("run_id", "")
    raw_data = audit_data.get("raw_data", {})
    ga4_raw_data = audit_data.get("ga4_raw_data", {})

    composite = scoring.get("composite_score", 0) or 0
    risk_band_label = scoring.get("risk_band", "")
    capital_impl = scoring.get("capital_implication", "")
    confidence = scoring.get("confidence", "")
    flags_count = scoring.get("red_flags_count", len(red_flags))

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 1: DASHBOARD
    # ═════════════════════════════════════════════════════════════════════════
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash.sheet_properties.tabColor = "1B2A4A"
    _set_col_widths(ws_dash, [5, 35, 18, 18, 18, 18, 25])

    _title_row(ws_dash, 1, "Media Integrity Engine — Audit Evaluation", 7)

    date_str = f"{date_range.get('start', '')} to {date_range.get('end', '')}"
    _title_row(ws_dash, 2,
               f"Client: {account_name}  |  Account: {account_id}  |  Period: {date_str}",
               7, font=WHITE_FONT_SM, fill=HEADER_BG)

    row = 4
    _title_row(ws_dash, row, "DOMAIN SCORES", 7, font=SUBTITLE_FONT, fill=DOMAIN_BG)
    row += 1

    _header_row(ws_dash, row, ["#", "Domain", "Score", "Risk Band", "Findings", "Weight", "Weighted"])
    row += 1

    idx = 0
    for key in DOMAIN_ORDER:
        idx += 1
        d = domain_scores.get(key, {})
        score_val = d.get("value", 0) or 0
        weight = DOMAIN_WEIGHTS.get(key, 0)
        label = DOMAIN_LABELS.get(key, key)
        band_name, band_fill = _risk_band(score_val)
        findings = d.get("key_findings", [])
        findings_str = "; ".join(findings) if findings else "—"
        weighted = round(score_val * weight / 100, 1)

        ws_dash.cell(row=row, column=1, value=idx).font = DARK_FONT_BOLD
        ws_dash.cell(row=row, column=1).alignment = CENTER
        ws_dash.cell(row=row, column=2, value=label).font = DARK_FONT_BOLD
        ws_dash.cell(row=row, column=2).alignment = LEFT_WRAP

        sc = ws_dash.cell(row=row, column=3, value=score_val)
        sc.font = Font(color="FFFFFF", bold=True, size=14)
        sc.fill = band_fill
        sc.alignment = CENTER

        ws_dash.cell(row=row, column=4, value=band_name).font = DARK_FONT_BOLD
        ws_dash.cell(row=row, column=4).alignment = CENTER

        ws_dash.cell(row=row, column=5, value=findings_str).font = DARK_FONT
        ws_dash.cell(row=row, column=5).alignment = LEFT_WRAP

        ws_dash.cell(row=row, column=6, value=f"{weight}%").font = DARK_FONT
        ws_dash.cell(row=row, column=6).alignment = CENTER

        ws_dash.cell(row=row, column=7, value=weighted).font = DARK_FONT_BOLD
        ws_dash.cell(row=row, column=7).alignment = CENTER

        alt_fill = LIGHT_GRAY_BG if idx % 2 == 0 else WHITE_BG
        _apply_row_fill(ws_dash, row, 1, 7, alt_fill, exclude={3})
        _apply_row_border(ws_dash, row, 1, 7)
        ws_dash.row_dimensions[row].height = 40
        row += 1

    # Composite Score
    row += 1
    _title_row(ws_dash, row, "COMPOSITE INTEGRITY SCORE", 7, font=SUBTITLE_FONT, fill=DOMAIN_BG)
    row += 1

    band_name, band_fill = _risk_band(composite)

    ws_dash.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws_dash.cell(row=row, column=1, value="Composite Score").font = DARK_FONT_BOLD
    ws_dash.cell(row=row, column=1).alignment = Alignment(horizontal="right", vertical="center")
    sc = ws_dash.cell(row=row, column=3, value=composite)
    sc.font = Font(color="FFFFFF", bold=True, size=18)
    sc.fill = band_fill
    sc.alignment = CENTER
    ws_dash.cell(row=row, column=4, value=band_name).font = DARK_FONT_BOLD
    ws_dash.cell(row=row, column=4).alignment = CENTER
    ws_dash.cell(row=row, column=5, value=f"Confidence: {confidence}").font = DARK_FONT
    ws_dash.cell(row=row, column=5).alignment = CENTER
    _apply_row_border(ws_dash, row, 1, 7)
    row += 1

    # Capital Implication
    ws_dash.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws_dash.cell(row=row, column=1, value="Capital Implication").font = DARK_FONT_BOLD
    ws_dash.cell(row=row, column=1).alignment = Alignment(horizontal="right", vertical="center")
    ci = ws_dash.cell(row=row, column=3, value=capital_impl)
    ci.font = Font(color="FFFFFF", bold=True, size=14)
    ci.alignment = CENTER
    impl_fills = {"REDUCE": RED_BG, "REWEIGHT": ORANGE_BG, "TEST": YELLOW_BG, "HOLD": GREEN_BG}
    ci.fill = impl_fills.get(capital_impl, YELLOW_BG)
    ws_dash.merge_cells(start_row=row, start_column=4, end_row=row, end_column=7)
    ws_dash.cell(row=row, column=4, value=IMPL_DESCRIPTIONS.get(capital_impl, "")).font = DARK_FONT
    ws_dash.cell(row=row, column=4).alignment = LEFT_WRAP
    _apply_row_border(ws_dash, row, 1, 7)
    row += 1

    # Red Flags Summary
    row += 1
    _title_row(ws_dash, row, "RED FLAGS SUMMARY", 7, font=SUBTITLE_FONT, fill=DOMAIN_BG)
    row += 1

    _header_row(ws_dash, row, ["", "Total Rules", "Triggered", "Passed", "Critical", "High", "Medium"])
    row += 1

    sev_counts = {"critical": 0, "high": 0, "medium": 0, "low": 0}
    for f in red_flags:
        sev = f.get("severity", "medium").lower()
        if sev in sev_counts:
            sev_counts[sev] += 1

    total_rules = scoring.get("total_rules_evaluated", 15)
    passed = total_rules - flags_count

    vals = ["", total_rules, flags_count, passed,
            sev_counts["critical"], sev_counts["high"], sev_counts["medium"]]
    for col, v in enumerate(vals, 1):
        c = ws_dash.cell(row=row, column=col, value=v)
        c.font = DARK_FONT_BOLD
        c.alignment = CENTER
        c.border = THIN_BORDER

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 2: RED FLAGS DETAIL
    # ═════════════════════════════════════════════════════════════════════════
    ws_flags = wb.create_sheet("Red Flags")
    ws_flags.sheet_properties.tabColor = "E74C3C"
    _set_col_widths(ws_flags, [10, 12, 30, 14, 55, 55])

    _title_row(ws_flags, 1, f"Red Flags — {account_name}", 6)
    _title_row(ws_flags, 2,
               f"{flags_count} red flag(s) triggered  |  Run: {run_id[:8] if run_id else ''}",
               6, font=WHITE_FONT_SM, fill=HEADER_BG)

    row = 4
    _header_row(ws_flags, row, ["Rule ID", "Severity", "Title", "Domain", "Description", "Recommendation"])
    row += 1

    if not red_flags:
        ws_flags.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws_flags.cell(row=row, column=1, value="No red flags detected.").font = DARK_FONT
        ws_flags.cell(row=row, column=1).alignment = CENTER
    else:
        current_domain = None
        for f in red_flags:
            domain = f.get("domain", "")
            domain_label = DOMAIN_LABELS.get(domain, domain)

            if domain != current_domain:
                current_domain = domain
                ws_flags.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                c = ws_flags.cell(row=row, column=1, value=f"  {domain_label}")
                c.font = WHITE_FONT
                c.fill = DOMAIN_BG
                c.alignment = Alignment(horizontal="left", vertical="center")
                for col in range(1, 7):
                    ws_flags.cell(row=row, column=col).fill = DOMAIN_BG
                row += 1

            severity = f.get("severity", "medium").lower()

            ws_flags.cell(row=row, column=1, value=f.get("id", f.get("rule_id_raw", ""))).font = DARK_FONT_BOLD
            ws_flags.cell(row=row, column=1).alignment = CENTER

            sev_cell = ws_flags.cell(row=row, column=2, value=severity.upper())
            sev_cell.font = SEVERITY_FONTS.get(severity, DARK_FONT_BOLD)
            sev_cell.fill = SEVERITY_FILLS.get(severity, YELLOW_BG)
            sev_cell.alignment = CENTER

            ws_flags.cell(row=row, column=3, value=f.get("title", "")).font = DARK_FONT_BOLD
            ws_flags.cell(row=row, column=3).alignment = LEFT_WRAP

            ws_flags.cell(row=row, column=4, value=domain_label).font = DARK_FONT
            ws_flags.cell(row=row, column=4).alignment = LEFT_WRAP

            ws_flags.cell(row=row, column=5, value=f.get("description", "")).font = DARK_FONT
            ws_flags.cell(row=row, column=5).alignment = LEFT_WRAP

            ws_flags.cell(row=row, column=6, value=f.get("recommendation", "")).font = DARK_FONT
            ws_flags.cell(row=row, column=6).alignment = LEFT_WRAP

            _apply_row_fill(ws_flags, row, 1, 6, FAIL_BG, exclude={2})
            _apply_row_border(ws_flags, row, 1, 6)
            ws_flags.row_dimensions[row].height = 50
            row += 1

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 3: DOMAIN BREAKDOWN
    # ═════════════════════════════════════════════════════════════════════════
    ws_domains = wb.create_sheet("Domain Breakdown")
    ws_domains.sheet_properties.tabColor = "2C3E6B"
    _set_col_widths(ws_domains, [30, 15, 15, 15, 60])

    _title_row(ws_domains, 1, f"Domain Score Breakdown — {account_name}", 5)

    row = 3
    for key in DOMAIN_ORDER:
        d = domain_scores.get(key, {})
        label = DOMAIN_LABELS.get(key, key)
        score_val = d.get("value", 0) or 0
        band_name, band_fill = _risk_band(score_val)
        sub_scores = d.get("sub_scores", {})
        completeness = d.get("data_completeness", 0)

        _title_row(ws_domains, row, f"{label}  —  {score_val}/100  ({band_name})", 5,
                   font=SUBTITLE_FONT, fill=DOMAIN_BG)
        row += 1

        if sub_scores:
            _header_row(ws_domains, row, ["Sub-metric", "Value", "", "", ""])
            row += 1
            for sk, sv in sub_scores.items():
                pretty_key = sk.replace("_", " ").title()
                ws_domains.cell(row=row, column=1, value=pretty_key).font = DARK_FONT
                ws_domains.cell(row=row, column=1).alignment = LEFT_WRAP

                if isinstance(sv, float) and sv <= 1.0:
                    ws_domains.cell(row=row, column=2, value=f"{sv:.1%}").font = DARK_FONT_BOLD
                else:
                    ws_domains.cell(row=row, column=2, value=str(sv)).font = DARK_FONT_BOLD
                ws_domains.cell(row=row, column=2).alignment = CENTER
                _apply_row_border(ws_domains, row, 1, 2)
                row += 1

        findings = d.get("key_findings", [])
        if findings:
            _header_row(ws_domains, row, ["Key Findings", "", "", "", ""])
            row += 1
            for finding in findings:
                ws_domains.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
                ws_domains.cell(row=row, column=1, value=f"• {finding}").font = DARK_FONT
                ws_domains.cell(row=row, column=1).alignment = LEFT_WRAP
                ws_domains.row_dimensions[row].height = 30
                row += 1

        ws_domains.cell(row=row, column=1, value="Data Completeness").font = DARK_FONT
        if isinstance(completeness, (int, float)):
            ws_domains.cell(row=row, column=2, value=f"{completeness:.0%}").font = DARK_FONT_BOLD
        else:
            ws_domains.cell(row=row, column=2, value=str(completeness)).font = DARK_FONT_BOLD
        ws_domains.cell(row=row, column=2).alignment = CENTER
        row += 2

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 4: RECOMMENDATIONS
    # ═════════════════════════════════════════════════════════════════════════
    ws_recs = wb.create_sheet("Recommendations")
    ws_recs.sheet_properties.tabColor = "F39C12"
    _set_col_widths(ws_recs, [5, 12, 22, 45, 45, 35])

    _title_row(ws_recs, 1, f"Top Recommendations — {account_name}", 6)

    severity_order = {"critical": 0, "high": 1, "medium": 2, "low": 3}
    sorted_flags = sorted(
        red_flags,
        key=lambda f: severity_order.get(f.get("severity", "medium").lower(), 99),
    )
    top_flags = sorted_flags[:10]

    _title_row(ws_recs, 2,
               f"{len(top_flags)} action item(s) prioritized by severity",
               6, font=WHITE_FONT_SM, fill=HEADER_BG)

    row = 4
    _header_row(ws_recs, row, ["#", "Priority", "Domain", "NOW (Current State)",
                                "SUGGESTION (Action Required)", "WHY (Impact)"])
    row += 1

    if not top_flags:
        ws_recs.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws_recs.cell(row=row, column=1,
                     value="No action items — all checks passed.").font = DARK_FONT
        ws_recs.cell(row=row, column=1).alignment = CENTER
    else:
        for idx, flag in enumerate(top_flags, 1):
            severity = flag.get("severity", "medium").lower()
            domain = flag.get("domain", "")
            domain_label = DOMAIN_LABELS.get(domain, domain)

            ws_recs.cell(row=row, column=1, value=idx).font = DARK_FONT_BOLD
            ws_recs.cell(row=row, column=1).alignment = CENTER

            sev_cell = ws_recs.cell(row=row, column=2, value=severity.upper())
            sev_cell.font = SEVERITY_FONTS.get(severity, DARK_FONT_BOLD)
            sev_cell.fill = SEVERITY_FILLS.get(severity, YELLOW_BG)
            sev_cell.alignment = CENTER

            ws_recs.cell(row=row, column=3, value=domain_label).font = DARK_FONT
            ws_recs.cell(row=row, column=3).alignment = LEFT_WRAP

            ws_recs.cell(row=row, column=4,
                         value=flag.get("description", "")).font = DARK_FONT
            ws_recs.cell(row=row, column=4).alignment = LEFT_WRAP

            ws_recs.cell(row=row, column=5,
                         value=flag.get("recommendation", "")).font = DARK_FONT
            ws_recs.cell(row=row, column=5).alignment = LEFT_WRAP

            evidence = flag.get("evidence", {})
            if isinstance(evidence, dict):
                why = "; ".join(f"{k}: {v}" for k, v in evidence.items()) if evidence else ""
            else:
                why = str(evidence) if evidence else ""
            if not why:
                why = flag.get("triggered_by", "")
            ws_recs.cell(row=row, column=6, value=why).font = DARK_FONT
            ws_recs.cell(row=row, column=6).alignment = LEFT_WRAP

            alt_fill = LIGHT_GRAY_BG if idx % 2 == 0 else WHITE_BG
            _apply_row_fill(ws_recs, row, 1, 6, alt_fill, exclude={2})
            _apply_row_border(ws_recs, row, 1, 6)
            ws_recs.row_dimensions[row].height = 60
            row += 1

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 5: IMPRESSION SHARE (only if raw_data available)
    # ═════════════════════════════════════════════════════════════════════════
    is_data = raw_data.get("impression_share", [])
    if is_data:
        ws_is = wb.create_sheet("Impression Share")
        ws_is.sheet_properties.tabColor = "3498DB"
        _set_col_widths(ws_is, [35, 18, 14, 14, 14, 16, 16])

        _title_row(ws_is, 1, f"Impression Share — {account_name}", 7)
        _title_row(ws_is, 2, f"Search impression share metrics by campaign  |  Period: {date_str}",
                   7, font=WHITE_FONT_SM, fill=HEADER_BG)

        row = 4
        _header_row(ws_is, row, ["Campaign", "Type", "Search IS %", "Top IS %",
                                  "Abs Top IS %", "Lost to Rank %", "Lost to Budget %"])
        row += 1

        # Aggregate daily rows by campaign (avg IS metrics)
        is_agg = {}
        for r in is_data:
            camp = r.get("campaign", {}) if isinstance(r.get("campaign"), dict) else {}
            metrics = r.get("metrics", {}) if isinstance(r.get("metrics"), dict) else r
            cname = camp.get("name", r.get("campaign_name", ""))
            if not cname:
                continue
            if cname not in is_agg:
                is_agg[cname] = {
                    "type": camp.get("advertising_channel_type", ""),
                    "vals": [], "count": 0,
                }
            is_agg[cname]["count"] += 1
            is_agg[cname]["vals"].append(metrics)

        is_metric_keys = [
            "search_impression_share", "search_top_impression_share",
            "search_absolute_top_impression_share",
            "search_rank_lost_impression_share", "search_budget_lost_impression_share",
        ]

        for idx, (cname, agg) in enumerate(sorted(is_agg.items()), 1):
            ws_is.cell(row=row, column=1, value=cname).font = DARK_FONT
            ws_is.cell(row=row, column=1).alignment = LEFT_WRAP
            ws_is.cell(row=row, column=2, value=agg["type"]).font = DARK_FONT
            ws_is.cell(row=row, column=2).alignment = CENTER

            for ki, key in enumerate(is_metric_keys):
                col = 3 + ki
                vals = [float(m.get(key, 0) or 0) for m in agg["vals"] if m.get(key) is not None]
                if vals:
                    avg = sum(vals) / len(vals)
                    display = f"{avg:.1%}" if avg <= 1.0 else f"{avg:.1f}%"
                else:
                    display = "—"
                c = ws_is.cell(row=row, column=col, value=display)
                c.font = DARK_FONT
                c.alignment = CENTER

            alt_fill = LIGHT_GRAY_BG if idx % 2 == 0 else WHITE_BG
            _apply_row_fill(ws_is, row, 1, 7, alt_fill)
            _apply_row_border(ws_is, row, 1, 7)
            row += 1

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 6: BUDGET & CAMPAIGNS (only if raw_data available)
    # ═════════════════════════════════════════════════════════════════════════
    campaign_data = raw_data.get("campaign_performance", [])
    if campaign_data:
        ws_budget = wb.create_sheet("Budget & Campaigns")
        ws_budget.sheet_properties.tabColor = "27AE60"
        _set_col_widths(ws_budget, [35, 18, 22, 14, 12, 14, 12, 12])

        _title_row(ws_budget, 1, f"Budget & Campaigns — {account_name}", 8)
        _title_row(ws_budget, 2,
                   f"Campaign performance and budget allocation  |  Period: {date_str}",
                   8, font=WHITE_FONT_SM, fill=HEADER_BG)

        # Build bidding strategy lookup (nested: row.campaign.name / .bidding_strategy_type)
        bidding_map = {}
        for br in raw_data.get("bidding_strategies", []):
            bc = br.get("campaign", {}) if isinstance(br.get("campaign"), dict) else {}
            bname = bc.get("name", br.get("campaign_name", ""))
            btype = bc.get("bidding_strategy_type", br.get("bidding_strategy_type", ""))
            if bname:
                bidding_map[bname] = btype

        # Aggregate daily rows by campaign (sum cost/conversions/value)
        camp_agg = {}
        for r in campaign_data:
            camp = r.get("campaign", {}) if isinstance(r.get("campaign"), dict) else {}
            metrics = r.get("metrics", {}) if isinstance(r.get("metrics"), dict) else r
            cname = camp.get("name", r.get("campaign_name", ""))
            if not cname:
                continue
            if cname not in camp_agg:
                camp_agg[cname] = {
                    "type": camp.get("advertising_channel_type", ""),
                    "cost": 0, "conversions": 0, "conv_value": 0,
                }
            cost_raw = metrics.get("cost_micros", metrics.get("cost", 0))
            cost_val = float(cost_raw or 0)
            # cost_micros is in micros (divide by 1e6)
            if cost_val > 1_000_000:
                cost_val = cost_val / 1_000_000
            camp_agg[cname]["cost"] += cost_val
            camp_agg[cname]["conversions"] += float(metrics.get("conversions", 0) or 0)
            camp_agg[cname]["conv_value"] += float(
                metrics.get("conversions_value", metrics.get("conversion_value", 0)) or 0
            )

        total_cost = sum(a["cost"] for a in camp_agg.values())

        row = 4
        _header_row(ws_budget, row, ["Campaign", "Type", "Bidding Strategy",
                                      "Cost", "% of Total", "Conversions", "CPA", "ROAS"])
        row += 1

        total_conv = 0
        total_value = 0
        sorted_campaigns = sorted(camp_agg.items(), key=lambda x: x[1]["cost"], reverse=True)

        for idx, (cname, agg) in enumerate(sorted_campaigns, 1):
            cost = agg["cost"]
            conv = agg["conversions"]
            conv_val = agg["conv_value"]
            total_conv += conv
            total_value += conv_val

            pct = cost / total_cost if total_cost > 0 else 0
            cpa = cost / conv if conv > 0 else None
            roas = conv_val / cost if cost > 0 else None
            bidding = bidding_map.get(cname, "—")

            ws_budget.cell(row=row, column=1, value=cname).font = DARK_FONT
            ws_budget.cell(row=row, column=1).alignment = LEFT_WRAP
            ws_budget.cell(row=row, column=2, value=agg["type"]).font = DARK_FONT
            ws_budget.cell(row=row, column=2).alignment = CENTER
            ws_budget.cell(row=row, column=3, value=bidding).font = DARK_FONT
            ws_budget.cell(row=row, column=3).alignment = LEFT_WRAP

            ws_budget.cell(row=row, column=4, value=round(cost, 2)).font = DARK_FONT
            ws_budget.cell(row=row, column=4).alignment = CENTER
            ws_budget.cell(row=row, column=4).number_format = '#,##0.00'

            ws_budget.cell(row=row, column=5, value=round(pct, 3)).font = DARK_FONT
            ws_budget.cell(row=row, column=5).alignment = CENTER
            ws_budget.cell(row=row, column=5).number_format = '0.0%'

            ws_budget.cell(row=row, column=6, value=round(conv, 1)).font = DARK_FONT
            ws_budget.cell(row=row, column=6).alignment = CENTER

            ws_budget.cell(row=row, column=7,
                           value=round(cpa, 2) if cpa is not None else "—").font = DARK_FONT
            ws_budget.cell(row=row, column=7).alignment = CENTER
            if cpa is not None:
                ws_budget.cell(row=row, column=7).number_format = '#,##0.00'

            ws_budget.cell(row=row, column=8,
                           value=round(roas, 2) if roas is not None else "—").font = DARK_FONT
            ws_budget.cell(row=row, column=8).alignment = CENTER

            alt_fill = LIGHT_GRAY_BG if idx % 2 == 0 else WHITE_BG
            _apply_row_fill(ws_budget, row, 1, 8, alt_fill)
            _apply_row_border(ws_budget, row, 1, 8)
            row += 1

        # Totals row
        total_cpa = total_cost / total_conv if total_conv > 0 else None
        total_roas = total_value / total_cost if total_cost > 0 else None

        ws_budget.cell(row=row, column=1, value="TOTAL").font = WHITE_FONT
        for col in range(1, 9):
            ws_budget.cell(row=row, column=col).fill = DOMAIN_BG
            ws_budget.cell(row=row, column=col).border = THIN_BORDER
        ws_budget.cell(row=row, column=4, value=round(total_cost, 2)).font = WHITE_FONT
        ws_budget.cell(row=row, column=4).alignment = CENTER
        ws_budget.cell(row=row, column=4).number_format = '#,##0.00'
        ws_budget.cell(row=row, column=5, value=1.0).font = WHITE_FONT
        ws_budget.cell(row=row, column=5).alignment = CENTER
        ws_budget.cell(row=row, column=5).number_format = '0.0%'
        ws_budget.cell(row=row, column=6, value=round(total_conv, 1)).font = WHITE_FONT
        ws_budget.cell(row=row, column=6).alignment = CENTER
        ws_budget.cell(row=row, column=7,
                       value=round(total_cpa, 2) if total_cpa else "—").font = WHITE_FONT
        ws_budget.cell(row=row, column=7).alignment = CENTER
        ws_budget.cell(row=row, column=8,
                       value=round(total_roas, 2) if total_roas else "—").font = WHITE_FONT
        ws_budget.cell(row=row, column=8).alignment = CENTER

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 7: CONVERSION SETUP (only if raw_data available)
    # ═════════════════════════════════════════════════════════════════════════
    conv_data = raw_data.get("conversion_actions", [])
    if conv_data:
        ws_conv = wb.create_sheet("Conversion Setup")
        ws_conv.sheet_properties.tabColor = "8E44AD"
        _set_col_widths(ws_conv, [35, 16, 20, 16, 16, 18, 16, 12])

        _title_row(ws_conv, 1, f"Conversion Setup — {account_name}", 8)
        _title_row(ws_conv, 2, f"{len(conv_data)} conversion action(s) configured",
                   8, font=WHITE_FONT_SM, fill=HEADER_BG)

        row = 4
        _header_row(ws_conv, row, ["Conversion Action", "Category", "Attribution Model",
                                    "Lookback (click)", "Lookback (view)", "Counting",
                                    "In Conversions", "Status"])
        row += 1

        dda_count = 0
        included_count = 0

        for idx, ca in enumerate(conv_data, 1):
            # Handle nested structure: row.conversion_action.{field}
            ca_inner = ca.get("conversion_action", {}) if isinstance(ca.get("conversion_action"), dict) else ca
            attr_settings = ca_inner.get("attribution_model_settings", {})
            if isinstance(attr_settings, dict):
                attr_model = attr_settings.get("attribution_model", "")
            else:
                attr_model = ca_inner.get("attribution_model", "")

            name = ca_inner.get("name", "")
            category = ca_inner.get("category", ca_inner.get("type_", ca_inner.get("type", "")))
            lb_click = ca_inner.get("click_through_lookback_window_days",
                                    ca_inner.get("lookback_window_days", ""))
            lb_view = ca_inner.get("view_through_lookback_window_days", "")
            counting = ca_inner.get("counting_type", "")
            included = ca_inner.get("include_in_conversions_metric",
                                    ca_inner.get("include_in_conversions",
                                                  ca_inner.get("primary_for_goal", "")))
            status = ca_inner.get("status", "")

            if "DATA_DRIVEN" in str(attr_model).upper():
                dda_count += 1
            if str(included).lower() in ("true", "yes", "1"):
                included_count += 1

            ws_conv.cell(row=row, column=1, value=name).font = DARK_FONT
            ws_conv.cell(row=row, column=1).alignment = LEFT_WRAP

            ws_conv.cell(row=row, column=2, value=str(category)).font = DARK_FONT
            ws_conv.cell(row=row, column=2).alignment = CENTER

            attr_cell = ws_conv.cell(row=row, column=3, value=str(attr_model))
            attr_cell.font = DARK_FONT_BOLD
            attr_cell.alignment = CENTER
            if "DATA_DRIVEN" in str(attr_model).upper():
                attr_cell.fill = PASS_BG
            elif "LAST_CLICK" in str(attr_model).upper():
                attr_cell.fill = FAIL_BG

            ws_conv.cell(row=row, column=4, value=str(lb_click)).font = DARK_FONT
            ws_conv.cell(row=row, column=4).alignment = CENTER

            ws_conv.cell(row=row, column=5, value=str(lb_view)).font = DARK_FONT
            ws_conv.cell(row=row, column=5).alignment = CENTER

            ws_conv.cell(row=row, column=6, value=str(counting)).font = DARK_FONT
            ws_conv.cell(row=row, column=6).alignment = CENTER

            incl_cell = ws_conv.cell(row=row, column=7, value=str(included))
            incl_cell.font = DARK_FONT
            incl_cell.alignment = CENTER
            if str(included).lower() in ("true", "yes", "1"):
                incl_cell.fill = PASS_BG

            ws_conv.cell(row=row, column=8, value=str(status)).font = DARK_FONT
            ws_conv.cell(row=row, column=8).alignment = CENTER

            alt_fill = LIGHT_GRAY_BG if idx % 2 == 0 else WHITE_BG
            _apply_row_fill(ws_conv, row, 1, 8, alt_fill, exclude={3, 7})
            _apply_row_border(ws_conv, row, 1, 8)
            row += 1

        # Summary
        row += 1
        _title_row(ws_conv, row, "Summary", 8, font=SUBTITLE_FONT, fill=DOMAIN_BG)
        row += 1
        dda_pct = dda_count / len(conv_data) if conv_data else 0
        summary_items = [
            f"Total conversion actions: {len(conv_data)}",
            f"Using Data-Driven Attribution: {dda_count} ({dda_pct:.0%})",
            f"Included in conversions column: {included_count}",
        ]
        for s in summary_items:
            ws_conv.cell(row=row, column=1, value=s).font = DARK_FONT
            ws_conv.cell(row=row, column=1).alignment = LEFT_WRAP
            row += 1

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 8: GA4 CHANNEL PERFORMANCE (only if ga4_raw_data available)
    # ═════════════════════════════════════════════════════════════════════════
    ga4_channel_data = ga4_raw_data.get("channel_revenue", [])
    if ga4_channel_data:
        ws_ga4ch = wb.create_sheet("GA4 Channel Performance")
        ws_ga4ch.sheet_properties.tabColor = "E67E22"
        _set_col_widths(ws_ga4ch, [28, 16, 16, 16, 14, 14, 14, 14])

        _title_row(ws_ga4ch, 1, f"GA4 Channel Performance — {account_name}", 8)
        _title_row(ws_ga4ch, 2,
                   f"Revenue and engagement by channel group  |  Period: {date_str}",
                   8, font=WHITE_FONT_SM, fill=HEADER_BG)

        # Aggregate daily rows by channel group
        ch_agg = {}
        for r in ga4_channel_data:
            ch = r.get("sessionDefaultChannelGroup", "(not set)")
            if ch not in ch_agg:
                ch_agg[ch] = {
                    "revenue": 0, "transactions": 0, "purchase_revenue": 0,
                    "sessions": 0, "engaged_sessions": 0, "users": 0, "conversions": 0,
                }
            ch_agg[ch]["revenue"] += float(r.get("totalRevenue", 0) or 0)
            ch_agg[ch]["transactions"] += int(float(r.get("transactions", 0) or 0))
            ch_agg[ch]["purchase_revenue"] += float(r.get("purchaseRevenue", 0) or 0)
            ch_agg[ch]["sessions"] += int(float(r.get("sessions", 0) or 0))
            ch_agg[ch]["engaged_sessions"] += int(float(r.get("engagedSessions", 0) or 0))
            ch_agg[ch]["users"] += int(float(r.get("totalUsers", 0) or 0))
            ch_agg[ch]["conversions"] += float(r.get("conversions", 0) or 0)

        total_rev = sum(a["revenue"] for a in ch_agg.values())
        total_sess = sum(a["sessions"] for a in ch_agg.values())

        row = 4
        _header_row(ws_ga4ch, row, ["Channel Group", "Revenue", "% of Revenue",
                                     "Transactions", "Sessions", "Engaged %",
                                     "Conversions", "Rev/Session"])
        row += 1

        sorted_channels = sorted(ch_agg.items(), key=lambda x: x[1]["revenue"], reverse=True)
        for idx, (ch_name, agg) in enumerate(sorted_channels, 1):
            rev = agg["revenue"]
            sess = agg["sessions"]
            eng = agg["engaged_sessions"]
            eng_rate = eng / sess if sess > 0 else 0
            rev_per_sess = rev / sess if sess > 0 else 0
            pct_rev = rev / total_rev if total_rev > 0 else 0

            ws_ga4ch.cell(row=row, column=1, value=ch_name).font = DARK_FONT
            ws_ga4ch.cell(row=row, column=1).alignment = LEFT_WRAP

            ws_ga4ch.cell(row=row, column=2, value=round(rev, 2)).font = DARK_FONT_BOLD
            ws_ga4ch.cell(row=row, column=2).alignment = CENTER
            ws_ga4ch.cell(row=row, column=2).number_format = '#,##0.00'

            ws_ga4ch.cell(row=row, column=3, value=round(pct_rev, 3)).font = DARK_FONT
            ws_ga4ch.cell(row=row, column=3).alignment = CENTER
            ws_ga4ch.cell(row=row, column=3).number_format = '0.0%'

            ws_ga4ch.cell(row=row, column=4, value=agg["transactions"]).font = DARK_FONT
            ws_ga4ch.cell(row=row, column=4).alignment = CENTER

            ws_ga4ch.cell(row=row, column=5, value=sess).font = DARK_FONT
            ws_ga4ch.cell(row=row, column=5).alignment = CENTER
            ws_ga4ch.cell(row=row, column=5).number_format = '#,##0'

            ws_ga4ch.cell(row=row, column=6, value=round(eng_rate, 3)).font = DARK_FONT
            ws_ga4ch.cell(row=row, column=6).alignment = CENTER
            ws_ga4ch.cell(row=row, column=6).number_format = '0.0%'

            ws_ga4ch.cell(row=row, column=7, value=round(agg["conversions"], 1)).font = DARK_FONT
            ws_ga4ch.cell(row=row, column=7).alignment = CENTER

            ws_ga4ch.cell(row=row, column=8, value=round(rev_per_sess, 2)).font = DARK_FONT
            ws_ga4ch.cell(row=row, column=8).alignment = CENTER
            ws_ga4ch.cell(row=row, column=8).number_format = '#,##0.00'

            alt_fill = LIGHT_GRAY_BG if idx % 2 == 0 else WHITE_BG
            _apply_row_fill(ws_ga4ch, row, 1, 8, alt_fill)
            _apply_row_border(ws_ga4ch, row, 1, 8)
            row += 1

        # Totals row
        ws_ga4ch.cell(row=row, column=1, value="TOTAL").font = WHITE_FONT
        for col in range(1, 9):
            ws_ga4ch.cell(row=row, column=col).fill = DOMAIN_BG
            ws_ga4ch.cell(row=row, column=col).border = THIN_BORDER
        ws_ga4ch.cell(row=row, column=2, value=round(total_rev, 2)).font = WHITE_FONT
        ws_ga4ch.cell(row=row, column=2).alignment = CENTER
        ws_ga4ch.cell(row=row, column=2).number_format = '#,##0.00'
        ws_ga4ch.cell(row=row, column=3, value=1.0).font = WHITE_FONT
        ws_ga4ch.cell(row=row, column=3).alignment = CENTER
        ws_ga4ch.cell(row=row, column=3).number_format = '0.0%'
        t_trans = sum(a["transactions"] for a in ch_agg.values())
        t_conv = sum(a["conversions"] for a in ch_agg.values())
        ws_ga4ch.cell(row=row, column=4, value=t_trans).font = WHITE_FONT
        ws_ga4ch.cell(row=row, column=4).alignment = CENTER
        ws_ga4ch.cell(row=row, column=5, value=total_sess).font = WHITE_FONT
        ws_ga4ch.cell(row=row, column=5).alignment = CENTER
        ws_ga4ch.cell(row=row, column=5).number_format = '#,##0'
        t_eng_rate = sum(a["engaged_sessions"] for a in ch_agg.values()) / total_sess if total_sess > 0 else 0
        ws_ga4ch.cell(row=row, column=6, value=round(t_eng_rate, 3)).font = WHITE_FONT
        ws_ga4ch.cell(row=row, column=6).alignment = CENTER
        ws_ga4ch.cell(row=row, column=6).number_format = '0.0%'
        ws_ga4ch.cell(row=row, column=7, value=round(t_conv, 1)).font = WHITE_FONT
        ws_ga4ch.cell(row=row, column=7).alignment = CENTER
        t_rps = total_rev / total_sess if total_sess > 0 else 0
        ws_ga4ch.cell(row=row, column=8, value=round(t_rps, 2)).font = WHITE_FONT
        ws_ga4ch.cell(row=row, column=8).alignment = CENTER
        ws_ga4ch.cell(row=row, column=8).number_format = '#,##0.00'

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 9: GA4 TRAFFIC ACQUISITION (only if ga4_raw_data available)
    # ═════════════════════════════════════════════════════════════════════════
    ga4_traffic_data = ga4_raw_data.get("traffic_acquisition", [])
    if ga4_traffic_data:
        ws_ga4tr = wb.create_sheet("GA4 Traffic Acquisition")
        ws_ga4tr.sheet_properties.tabColor = "3498DB"
        _set_col_widths(ws_ga4tr, [22, 16, 30, 14, 14, 14, 14, 14])

        _title_row(ws_ga4tr, 1, f"GA4 Traffic Acquisition — {account_name}", 8)
        _title_row(ws_ga4tr, 2,
                   f"Sessions by source / medium / campaign  |  Period: {date_str}",
                   8, font=WHITE_FONT_SM, fill=HEADER_BG)

        # Aggregate daily rows by source/medium/campaign
        tr_agg = {}
        for r in ga4_traffic_data:
            source = r.get("sessionSource", "(direct)")
            medium = r.get("sessionMedium", "(none)")
            campaign = r.get("sessionCampaignName", "(not set)")
            key = (source, medium, campaign)
            if key not in tr_agg:
                tr_agg[key] = {
                    "sessions": 0, "engaged_sessions": 0,
                    "conversions": 0, "revenue": 0,
                    "bounce_sum": 0, "bounce_count": 0,
                }
            tr_agg[key]["sessions"] += int(float(r.get("sessions", 0) or 0))
            tr_agg[key]["engaged_sessions"] += int(float(r.get("engagedSessions", 0) or 0))
            tr_agg[key]["conversions"] += float(r.get("conversions", 0) or 0)
            tr_agg[key]["revenue"] += float(r.get("totalRevenue", 0) or 0)
            br = float(r.get("bounceRate", 0) or 0)
            if br > 0:
                tr_agg[key]["bounce_sum"] += br
                tr_agg[key]["bounce_count"] += 1

        row = 4
        _header_row(ws_ga4tr, row, ["Source", "Medium", "Campaign",
                                     "Sessions", "Engaged %", "Bounce Rate",
                                     "Conversions", "Revenue"])
        row += 1

        sorted_traffic = sorted(tr_agg.items(), key=lambda x: x[1]["sessions"], reverse=True)
        # Limit to top 50 to keep the sheet readable
        for idx, ((source, medium, campaign), agg) in enumerate(sorted_traffic[:50], 1):
            sess = agg["sessions"]
            eng_rate = agg["engaged_sessions"] / sess if sess > 0 else 0
            avg_bounce = agg["bounce_sum"] / agg["bounce_count"] if agg["bounce_count"] > 0 else 0

            ws_ga4tr.cell(row=row, column=1, value=source).font = DARK_FONT
            ws_ga4tr.cell(row=row, column=1).alignment = LEFT_WRAP

            ws_ga4tr.cell(row=row, column=2, value=medium).font = DARK_FONT
            ws_ga4tr.cell(row=row, column=2).alignment = CENTER

            ws_ga4tr.cell(row=row, column=3, value=campaign).font = DARK_FONT
            ws_ga4tr.cell(row=row, column=3).alignment = LEFT_WRAP

            ws_ga4tr.cell(row=row, column=4, value=sess).font = DARK_FONT
            ws_ga4tr.cell(row=row, column=4).alignment = CENTER
            ws_ga4tr.cell(row=row, column=4).number_format = '#,##0'

            ws_ga4tr.cell(row=row, column=5, value=round(eng_rate, 3)).font = DARK_FONT
            ws_ga4tr.cell(row=row, column=5).alignment = CENTER
            ws_ga4tr.cell(row=row, column=5).number_format = '0.0%'

            ws_ga4tr.cell(row=row, column=6, value=round(avg_bounce, 3)).font = DARK_FONT
            ws_ga4tr.cell(row=row, column=6).alignment = CENTER
            ws_ga4tr.cell(row=row, column=6).number_format = '0.0%'

            ws_ga4tr.cell(row=row, column=7, value=round(agg["conversions"], 1)).font = DARK_FONT
            ws_ga4tr.cell(row=row, column=7).alignment = CENTER

            ws_ga4tr.cell(row=row, column=8, value=round(agg["revenue"], 2)).font = DARK_FONT
            ws_ga4tr.cell(row=row, column=8).alignment = CENTER
            ws_ga4tr.cell(row=row, column=8).number_format = '#,##0.00'

            alt_fill = LIGHT_GRAY_BG if idx % 2 == 0 else WHITE_BG
            _apply_row_fill(ws_ga4tr, row, 1, 8, alt_fill)
            _apply_row_border(ws_ga4tr, row, 1, 8)
            row += 1

        if len(sorted_traffic) > 50:
            row += 1
            ws_ga4tr.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            ws_ga4tr.cell(row=row, column=1,
                          value=f"Showing top 50 of {len(sorted_traffic)} source/medium/campaign combinations.").font = DARK_FONT
            ws_ga4tr.cell(row=row, column=1).alignment = CENTER

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 10: GA4 PAID VS ORGANIC (only if ga4_raw_data available)
    # ═════════════════════════════════════════════════════════════════════════
    ga4_pvo_data = ga4_raw_data.get("paid_vs_organic", [])
    if ga4_pvo_data:
        ws_pvo = wb.create_sheet("GA4 Paid vs Organic")
        ws_pvo.sheet_properties.tabColor = "27AE60"
        _set_col_widths(ws_pvo, [28, 14, 16, 16, 14, 14, 14])

        _title_row(ws_pvo, 1, f"GA4 Paid vs Organic — {account_name}", 7)
        _title_row(ws_pvo, 2,
                   f"Revenue and traffic split by paid / organic / other  |  Period: {date_str}",
                   7, font=WHITE_FONT_SM, fill=HEADER_BG)

        total_pvo_rev = sum(float(r.get("totalRevenue", 0) or 0) for r in ga4_pvo_data)
        total_pvo_sess = sum(int(float(r.get("sessions", 0) or 0)) for r in ga4_pvo_data)

        # ── Detail by channel group ──
        row = 4
        _header_row(ws_pvo, row, ["Channel Group", "Category", "Revenue",
                                   "% of Revenue", "Transactions", "Sessions", "Conversions"])
        row += 1

        sorted_pvo = sorted(ga4_pvo_data, key=lambda r: float(r.get("totalRevenue", 0) or 0), reverse=True)
        for idx, r in enumerate(sorted_pvo, 1):
            ch = r.get("sessionDefaultChannelGroup", "(not set)")
            cat = r.get("category", "other")
            rev = float(r.get("totalRevenue", 0) or 0)
            trans = int(float(r.get("transactions", 0) or 0))
            sess = int(float(r.get("sessions", 0) or 0))
            conv = float(r.get("conversions", 0) or 0)
            pct = rev / total_pvo_rev if total_pvo_rev > 0 else 0

            ws_pvo.cell(row=row, column=1, value=ch).font = DARK_FONT
            ws_pvo.cell(row=row, column=1).alignment = LEFT_WRAP

            cat_cell = ws_pvo.cell(row=row, column=2, value=cat.upper())
            cat_cell.font = DARK_FONT_BOLD
            cat_cell.alignment = CENTER
            if cat == "paid":
                cat_cell.fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
            elif cat == "organic":
                cat_cell.fill = PASS_BG

            ws_pvo.cell(row=row, column=3, value=round(rev, 2)).font = DARK_FONT
            ws_pvo.cell(row=row, column=3).alignment = CENTER
            ws_pvo.cell(row=row, column=3).number_format = '#,##0.00'

            ws_pvo.cell(row=row, column=4, value=round(pct, 3)).font = DARK_FONT
            ws_pvo.cell(row=row, column=4).alignment = CENTER
            ws_pvo.cell(row=row, column=4).number_format = '0.0%'

            ws_pvo.cell(row=row, column=5, value=trans).font = DARK_FONT
            ws_pvo.cell(row=row, column=5).alignment = CENTER

            ws_pvo.cell(row=row, column=6, value=sess).font = DARK_FONT
            ws_pvo.cell(row=row, column=6).alignment = CENTER
            ws_pvo.cell(row=row, column=6).number_format = '#,##0'

            ws_pvo.cell(row=row, column=7, value=round(conv, 1)).font = DARK_FONT
            ws_pvo.cell(row=row, column=7).alignment = CENTER

            alt_fill = LIGHT_GRAY_BG if idx % 2 == 0 else WHITE_BG
            _apply_row_fill(ws_pvo, row, 1, 7, alt_fill, exclude={2})
            _apply_row_border(ws_pvo, row, 1, 7)
            row += 1

        # ── Category summary ──
        row += 1
        _title_row(ws_pvo, row, "CATEGORY SUMMARY", 7, font=SUBTITLE_FONT, fill=DOMAIN_BG)
        row += 1
        _header_row(ws_pvo, row, ["Category", "", "Revenue", "% of Revenue",
                                   "Transactions", "Sessions", "Conversions"])
        row += 1

        cat_agg = {"paid": {"rev": 0, "trans": 0, "sess": 0, "conv": 0},
                   "organic": {"rev": 0, "trans": 0, "sess": 0, "conv": 0},
                   "other": {"rev": 0, "trans": 0, "sess": 0, "conv": 0}}
        for r in ga4_pvo_data:
            cat = r.get("category", "other")
            if cat not in cat_agg:
                cat = "other"
            cat_agg[cat]["rev"] += float(r.get("totalRevenue", 0) or 0)
            cat_agg[cat]["trans"] += int(float(r.get("transactions", 0) or 0))
            cat_agg[cat]["sess"] += int(float(r.get("sessions", 0) or 0))
            cat_agg[cat]["conv"] += float(r.get("conversions", 0) or 0)

        cat_fills = {
            "paid": PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"),
            "organic": PASS_BG,
            "other": LIGHT_GRAY_BG,
        }

        for cat_name in ("paid", "organic", "other"):
            ca = cat_agg[cat_name]
            pct = ca["rev"] / total_pvo_rev if total_pvo_rev > 0 else 0

            c1 = ws_pvo.cell(row=row, column=1, value=cat_name.upper())
            c1.font = DARK_FONT_BOLD
            c1.alignment = CENTER
            c1.fill = cat_fills.get(cat_name, WHITE_BG)

            ws_pvo.cell(row=row, column=3, value=round(ca["rev"], 2)).font = DARK_FONT_BOLD
            ws_pvo.cell(row=row, column=3).alignment = CENTER
            ws_pvo.cell(row=row, column=3).number_format = '#,##0.00'

            ws_pvo.cell(row=row, column=4, value=round(pct, 3)).font = DARK_FONT_BOLD
            ws_pvo.cell(row=row, column=4).alignment = CENTER
            ws_pvo.cell(row=row, column=4).number_format = '0.0%'

            ws_pvo.cell(row=row, column=5, value=ca["trans"]).font = DARK_FONT_BOLD
            ws_pvo.cell(row=row, column=5).alignment = CENTER

            ws_pvo.cell(row=row, column=6, value=ca["sess"]).font = DARK_FONT_BOLD
            ws_pvo.cell(row=row, column=6).alignment = CENTER
            ws_pvo.cell(row=row, column=6).number_format = '#,##0'

            ws_pvo.cell(row=row, column=7, value=round(ca["conv"], 1)).font = DARK_FONT_BOLD
            ws_pvo.cell(row=row, column=7).alignment = CENTER

            _apply_row_border(ws_pvo, row, 1, 7)
            row += 1

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 11: GA4 ATTRIBUTION COMPARISON (only if both GA4 + GAds data)
    # ═════════════════════════════════════════════════════════════════════════
    ga4_attr_data = ga4_raw_data.get("attribution", [])
    if ga4_attr_data and conv_data:
        ws_attr = wb.create_sheet("GA4 vs Ads Attribution")
        ws_attr.sheet_properties.tabColor = "8E44AD"
        _set_col_widths(ws_attr, [28, 18, 18, 18, 18, 16])

        _title_row(ws_attr, 1, f"GA4 vs Google Ads Attribution — {account_name}", 6)
        _title_row(ws_attr, 2,
                   f"GA4 DDA conversions by channel vs Google Ads reported conversions  |  Period: {date_str}",
                   6, font=WHITE_FONT_SM, fill=HEADER_BG)

        # GA4 attribution: channel group → conversions, revenue
        row = 4
        _title_row(ws_attr, row, "GA4 Attribution (Data-Driven)", 6,
                   font=SUBTITLE_FONT, fill=DOMAIN_BG)
        row += 1
        _header_row(ws_attr, row, ["Channel Group", "GA4 Conversions", "GA4 Revenue",
                                    "% of Conversions", "% of Revenue", ""])
        row += 1

        total_ga4_conv = sum(float(r.get("conversions", 0) or 0) for r in ga4_attr_data)
        total_ga4_rev = sum(float(r.get("totalRevenue", 0) or 0) for r in ga4_attr_data)

        sorted_attr = sorted(ga4_attr_data,
                             key=lambda r: float(r.get("totalRevenue", 0) or 0), reverse=True)
        for idx, r in enumerate(sorted_attr, 1):
            ch = r.get("sessionDefaultChannelGroup", "(not set)")
            conv = float(r.get("conversions", 0) or 0)
            rev = float(r.get("totalRevenue", 0) or 0)
            pct_conv = conv / total_ga4_conv if total_ga4_conv > 0 else 0
            pct_rev = rev / total_ga4_rev if total_ga4_rev > 0 else 0

            ws_attr.cell(row=row, column=1, value=ch).font = DARK_FONT
            ws_attr.cell(row=row, column=1).alignment = LEFT_WRAP

            ws_attr.cell(row=row, column=2, value=round(conv, 1)).font = DARK_FONT
            ws_attr.cell(row=row, column=2).alignment = CENTER

            ws_attr.cell(row=row, column=3, value=round(rev, 2)).font = DARK_FONT
            ws_attr.cell(row=row, column=3).alignment = CENTER
            ws_attr.cell(row=row, column=3).number_format = '#,##0.00'

            ws_attr.cell(row=row, column=4, value=round(pct_conv, 3)).font = DARK_FONT
            ws_attr.cell(row=row, column=4).alignment = CENTER
            ws_attr.cell(row=row, column=4).number_format = '0.0%'

            ws_attr.cell(row=row, column=5, value=round(pct_rev, 3)).font = DARK_FONT
            ws_attr.cell(row=row, column=5).alignment = CENTER
            ws_attr.cell(row=row, column=5).number_format = '0.0%'

            alt_fill = LIGHT_GRAY_BG if idx % 2 == 0 else WHITE_BG
            _apply_row_fill(ws_attr, row, 1, 6, alt_fill)
            _apply_row_border(ws_attr, row, 1, 6)
            row += 1

        # Totals
        ws_attr.cell(row=row, column=1, value="TOTAL").font = WHITE_FONT
        for col in range(1, 7):
            ws_attr.cell(row=row, column=col).fill = DOMAIN_BG
            ws_attr.cell(row=row, column=col).border = THIN_BORDER
        ws_attr.cell(row=row, column=2, value=round(total_ga4_conv, 1)).font = WHITE_FONT
        ws_attr.cell(row=row, column=2).alignment = CENTER
        ws_attr.cell(row=row, column=3, value=round(total_ga4_rev, 2)).font = WHITE_FONT
        ws_attr.cell(row=row, column=3).alignment = CENTER
        ws_attr.cell(row=row, column=3).number_format = '#,##0.00'
        row += 2

        # Cross-reference summary: GA4 total vs Google Ads total
        _title_row(ws_attr, row, "Cross-Platform Comparison", 6,
                   font=SUBTITLE_FONT, fill=DOMAIN_BG)
        row += 1

        # Compute Google Ads totals from conversion_actions
        gads_total_conv = 0
        for ca in conv_data:
            ca_inner = ca.get("conversion_action", {}) if isinstance(ca.get("conversion_action"), dict) else ca
            metrics = ca.get("metrics", ca_inner)
            gads_total_conv += float(metrics.get("all_conversions", metrics.get("conversions", 0)) or 0)

        comparison_items = [
            ("GA4 Total Conversions (DDA)", round(total_ga4_conv, 1)),
            ("GA4 Total Revenue", f"${total_ga4_rev:,.2f}"),
            ("Google Ads Total Conversions", round(gads_total_conv, 1) if gads_total_conv else "N/A"),
        ]
        if total_ga4_conv > 0 and gads_total_conv > 0:
            conv_delta = ((total_ga4_conv - gads_total_conv) / gads_total_conv) * 100
            comparison_items.append(("Conversion Delta (GA4 vs Ads)", f"{conv_delta:+.1f}%"))

        for label, value in comparison_items:
            ws_attr.cell(row=row, column=1, value=label).font = DARK_FONT_BOLD
            ws_attr.cell(row=row, column=1).alignment = LEFT_WRAP
            c = ws_attr.cell(row=row, column=2, value=str(value))
            c.font = DARK_FONT_BOLD
            c.alignment = CENTER
            _apply_row_border(ws_attr, row, 1, 2)
            row += 1

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET: ADVANCED HEALTH (Phase 3 metrics)
    # ═════════════════════════════════════════════════════════════════════════
    _build_advanced_health_sheet(wb, raw_data, account_name, date_str)

    # ═════════════════════════════════════════════════════════════════════════
    # LAST SHEET: EXECUTION LOG (always last)
    # ═════════════════════════════════════════════════════════════════════════
    ws_exec = wb.create_sheet("Execution")
    ws_exec.sheet_properties.tabColor = "27AE60"
    _set_col_widths(ws_exec, [25, 40])

    _title_row(ws_exec, 1, "Execution Details", 2)

    execution = audit_data.get("execution", {})
    row = 3
    meta_items = [
        ("Run ID", run_id),
        ("Account Name", account_name),
        ("Account ID", account_id),
        ("Date Range", date_str),
        ("Composite Score", composite),
        ("Risk Band", risk_band_label),
        ("Capital Implication", capital_impl),
        ("Confidence", confidence),
        ("Red Flags Count", flags_count),
        ("Duration (s)", execution.get("duration_seconds", "")),
        ("Source", execution.get("source", "")),
        ("Timestamp", execution.get("timestamp", "")),
    ]

    for label, value in meta_items:
        ws_exec.cell(row=row, column=1, value=label).font = DARK_FONT_BOLD
        ws_exec.cell(row=row, column=1).alignment = LEFT_WRAP
        ws_exec.cell(row=row, column=2, value=str(value) if value else "—").font = DARK_FONT
        ws_exec.cell(row=row, column=2).alignment = LEFT_WRAP
        _apply_row_border(ws_exec, row, 1, 2)
        row += 1

    # ═════════════════════════════════════════════════════════════════════════
    # SAVE TO BYTES
    # ═════════════════════════════════════════════════════════════════════════
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
